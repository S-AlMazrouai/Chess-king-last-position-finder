import requests
import json
import chess
import chess.pgn
import io
from collections import Counter
from openpyxl import load_workbook
import numpy

#API link: https://api.chess.com/pub/player/{user}/games/{year}/{month}/pgn
baseUrl='https://api.chess.com/pub/player/'

users=['Mazrouai']                                                                                          # You can add one or more chess.com profile/s, make sure to type the prfile name/s as it's/they're written in chess.com.

for user in users:                          
    years = range(2000,2022)                                                                                # Add the range of the years you want this code to analyze (from,to).
    months = ['01','02','03','04','05','06','07','08','09','10','11','12']                                  # Keep this as it is.
    count=0                         

    winBlackKingPos=[]                                                                                      # Array to collect King position in the games won as black.
    lossBlackKingPos=[]                                                                                     # Array to collect King position in the games lost as black.
    winWhiteKingPos=[]                                                                                      # Array to collect King position in the games won as white.
    lossWhiteKingPos=[]                                                                                     # Array to collect King position in the games lost as white.

    for i in years:                                                                                         # For loop to irritate through the specified years range.
        for j in months:                                                                                    # For loop to irritate through the monthes of the specified years.
            extension=str(str(user)+'/games/'+str(i)+'/'+str(j)+'/pgn')                                     # Creates the extension for the baseUrl.
            url=baseUrl+extension                                                                           # Merges baseUrl with the extension.
            response = requests.get(url)                            
            pgns = io.StringIO(response.text)                           
            if response.text == '':                                                                         # Checks if pgn file is empty and if it is, it jumps to the next PGN file.
                continue                            
            while True:                         
                games=chess.pgn.read_game(pgns)                                                             # Reads PGN file.
                if games == None:                                                                           # Checks if there is a game available to read inside the pgn file, if not it exits this loop to the next PGN file.
                    break                           
                if games.headers['Black'] == '?':                                                           # Checks if game data is missing, if true it jumps to the next game.
                    continue                            
                if games.headers['White'] == '?':                                                           # Checks if game data is missing, if true it jumps to the next game.
                    continue                            
                board=games.board()                         

                for move in games.mainline_moves():                                                         # Moves to the last position in the game.
                    board.push(move)                            

                map=board.piece_map()                                                                       # Collect the position of the pieces in thier last move.

                if games.headers['Black']== str(user):                                                      # Checks if the specified user is playing as black
                    for x,y in map.items():                         
                        if str(y) == 'k':                           
                            kingPos=chess.square_name(x)                                                    # Gets the black king postion.

                    if games.headers['Result'] == '0-1':                                                    # Collects the king position in the games won as black.
                        winBlackKingPos.append(kingPos)                         
                    if games.headers['Result'] == '1-0':                                                    # Collects the king position in the games lost as black.
                        lossBlackKingPos.append(kingPos)                            

                else:                                                                                       # If the if condition is not satisfied then the specificed user is playing as white.
                    for x,y in map.items():                         
                        if str(y) == 'K':                           
                            kingPos=chess.square_name(x)                                                    # Gets the white king postion.

                    if games.headers['Result'] == '0-1':                                                    # Collects the king position in the games lost as white.
                        lossWhiteKingPos.append(kingPos)                            
                    if games.headers['Result'] == '1-0':                                                    # Collects the king position in the games won as white.
                        winWhiteKingPos.append(kingPos)                         


    gamesWon=len(winBlackKingPos)+len(winWhiteKingPos)                                                      # Counts # of won games.
    gamesLost=len(lossBlackKingPos)+len(lossWhiteKingPos)                                                   # Counts # of lost games.
    gamesPlayed=gamesWon+gamesLost                                                                          # counts # of analyzed games

    print("Player: ",user)                                                                                  # Prints the name of the player.
    print("games played: ",gamesPlayed)                                                                     # Prints # of won games.
    print("games won: ",gamesWon)                                                                           # Prints # of lost games.
    print("games lost: ",gamesLost)                                                                         # Prints # of analyzed games
    print("\n")


    winWhiteKingPosCount= Counter(winWhiteKingPos)                                                          # Creates a list with a position and the number of times the wining white king was in that position.
    lossWhiteKingPosCount= Counter(lossWhiteKingPos)                                                        # Creates a list with a position and the number of times the losing white king was in that position.
    winBlackKingPosCount= Counter(winBlackKingPos)                                                          # Creates a list with a position and the number of times the wining black king was in that position.
    lossBlackKingPosCount= Counter(lossBlackKingPos)                                                        # Creates a list with a position and the number of times the losing black king was in that position.
    posCounts=[winWhiteKingPosCount,lossWhiteKingPosCount,winBlackKingPosCount,lossBlackKingPosCount]       # Merges the lists into an array.



    Data = load_workbook(filename='Data_Template.xlsx')                                                     # Opens the template excel file .
    sheets=Data.sheetnames                                                                                  # Register the sheets name.

    cellLetters=[]                                                                                          # Array for the cell letters in the excel file.
    cellNum=[]                                                                                              # Array for the cell numbers in the excel file.

    for j in range(8):                                                                                      # Generates cell letters to get the cells this code will work .
        for i in range(66, 74):
            cellLetters.append(chr(i))

    for i in [10,9,8,7,6,5,4,3]:                                                                            # Generates cell numbers to get the cells this code will work .
        for j in range(8):
            cellNum.append(i)

    c = 0                                                                                                   # This variable will be used as an index to go thorugh the lists that have been merged into an array.
    for sheet in sheets:                                                                                    # For loop to irritate through the excel sheets.
        
        workSheet=Data[sheet]
        posCount=posCounts[c]                                                                               # Gets the postion list.
        c=c+1
        
        for i in range(64):                                                                                 # For loop to go through the sheet cells and assign them the king recurrence value. 
            
            cell=str(cellLetters[i])+str(cellNum[i])                                                        # Constructs the excel cell name (e.g. A12).
            count=posCount[chess.square_name(i)]                                                            # Gets the king postion count that correlates with the cell name.
            if count== 0:                                                                                   # If king recurrence equals 0 set the cell to None.
                count= None
            workSheet[cell] = count                                                                         # Makes the cell value equales the king recurrence in that position.


    Data.save(filename='Data_'+str(user)+'.xlsx')                                                           # Saves the data into a new xlsx file
